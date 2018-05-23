#!/usr/bin/python3

"""
Author: P Surya Teja
Last Modified Date: 23-05-2018(DD-MM-YYYY)

The below script takes
1) the path of a valid excel file and
2) the path of an existing or new sqlite3 database
from the user and transfers the table in the active sheet
with sheet's title as the database table name.

If there are any repeating column names, the user will be prompted
for a new name for that column in the middle of execution.

The source excel file will not be modified.

If you are giving the name of existing database, make sure that it
has no table with sheet's title in it.
"""

import datetime
import openpyxl
import os
import pprint
import re
import sqlite3
import sys
from collections import OrderedDict

headings_types = {
    datetime.date: 'date',
    datetime.datetime: 'datetime',
    datetime.time: 'time',
    float: 'real',
    int: 'int',
    str: 'text'
}


def main():
    # prompting user for excel file path
    while True:
        excel_filename = input("Enter the absolute path of a valid excel file."
                               "\nLike 1)/home/user/example.xlsx in "
                               "mac or linux and\n"
                               "2) C:\\Users\\example.xlsx in windows\n>>>")
        if not os.path.exists(excel_filename):
            print("{} doesn't exist.".format(excel_filename))
        else:
            break

    # prompting user for database file path
    while True:
        db_filename = input("Enter the path of an existing sqlite3"
                            " database file or a new sqlite3 file.\n>>>")
        # opening the db file and establishing connection
        try:
            conn = sqlite3.connect(db_filename)
            break
        except sqlite3.OperationalError:
            print("Unable to open database. Permission denied.\n")
            continue
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    # removing empty rows and columns
    remove_empty(sheet)

    # headings is a dictionary with keys as table headings and
    # values as their SQL types
    headings = get_headings(sheet)  # type: dict
    table_name = "\"" + slugify(sheet.title) + "\""
    cursor = conn.cursor()

    # creating the table
    cursor.execute("drop table if exists {}".format(table_name))
    cursor.execute(create_table(headings, table_name))
    insert_values(table_name, cursor, sheet)
    conn.commit()

    print("All data from active sheet successfully written to the database.")

    # closing opened things
    conn.close()
    workbook.close()


def remove_empty(sheet):
    """
    Removes all rows and columns that are completely empty

    TODO: remove repeating code
    """
    all_rows = list(sheet.rows)
    row_idx = 2
    while row_idx <= sheet.max_row:
        row_values = list(set([cell.value
                               for cell in all_rows[row_idx - 1]]))
        if len(row_values) == 1 and row_values[0] is None:
            sheet.delete_rows(row_idx)
        else:
            row_idx += 1
    all_cols = list(sheet.columns)
    col_idx = 1
    while col_idx <= sheet.max_column:
        column_values = list(set([cell.value
                                  for cell in all_cols[col_idx - 1]]))
        if len(column_values) == 1 and column_values[0] is None:
            sheet.delete_cols(col_idx)
        else:
            col_idx += 1
    print(sheet.max_row)


def insert_values(table_name, cursor, sheet):
    """
    inserts values from row 2 to last row into database
    """
    all_rows = list(sheet.rows)
    placeholder = ", ".join('?' * sheet.max_column)
    for row in all_rows[1:]:
        try:
            cursor.execute(
                "insert into {} values ({})".
                format(table_name, placeholder), [cell.value for cell in row])
        except sqlite3.OperationalError:
            pprint.pprint([cell.value for cell in row])
            sys.exit(1)


def create_table(headings, table_name):
    """
    creates table in the database using keys and values of headings dict
    as column names and column types
    """
    temp = []
    for key, value in headings.items():
        temp.append("\"{}\" {}".format(key, value))
    query = ['create', 'table', table_name, '(', ", ".join(temp), ')']
    # print(" ".join(query))
    return " ".join(query)


def get_headings(sheet):
    """
    takes a worksheet object and returns a dictionary with keys as column
    names and their corresponding sql types as values
    """
    headings = OrderedDict()
    row = 1  # that is where headings live
    for col in range(1, sheet.max_column + 1):
        column_name = sheet.cell(row, col).value
        if column_name in (None, ""):
            column_name = "column" + str(col)
        column_name = slugify(column_name)
        column_name = check_repetitions(column_name,
                                        headings.keys(),
                                        sheet.cell(row, col).column)
        headings[column_name] = headings_types[type(sheet.cell(2, col).value)]
    return headings


def check_repetitions(column_name, headings_names, column):
    """
    If the table headings are repeated,
    it prompts user for new non repeating name
    """
    while True:
        if column_name in headings_names:
            column_name = input("column name {} of Column \'{}\'"
                                " already exists."
                                " Enter a new name: ".format(column_name,
                                                             column))
        else:
            break
    return column_name


def slugify(some_string):
    """
    allows only certain characters in the heading name,
    removes all others
    replaces multiple consequent spaces with a single underscore
    """
    sanity_regex = re.compile(r"""[^\w\s/%.'"()]+""")
    some_string = sanity_regex.sub('', some_string)
    space_regex = re.compile(r"\s+")
    return space_regex.sub("_", some_string.strip())


if __name__ == '__main__':
    main()
